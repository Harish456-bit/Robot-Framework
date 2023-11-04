from openpyxl import load_workbook
from openpyxl import load_workbook, Workbook

def get_excel_data(sheet_name, cell_address):
    try:
        workbook = load_workbook(filename="./Testdata/AutomationScript.xlsx")
        sheet = workbook[sheet_name]
        cell_value = sheet[cell_address].value
        return cell_value
    except Exception as e:
        return str(e)


def write_excel_data(sheet_name, cell_address, data):
    try:
        workbook = load_workbook(filename="./Testdata/AutomationScript.xlsx")
        sheet = workbook[sheet_name]
        sheet[cell_address] = data
        workbook.save("./Testdata/AutomationScript.xlsx")
        return "Data written successfully."
    except Exception as e:
        return str(e)